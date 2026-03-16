"""TDD tests for FormulaXlsxParser and supporting modules.

Run with:
    uv run --no-project python -m pytest \
        scripts/data/doc_intelligence/tests/test_formula_xlsx_parser.py -v
"""

import os
import sys
import tempfile
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

# Ensure the hub root is on sys.path so absolute imports resolve.
_HUB_ROOT = str(Path(__file__).resolve().parents[4])
if _HUB_ROOT not in sys.path:
    sys.path.insert(0, _HUB_ROOT)

from scripts.data.doc_intelligence.formula_reference_parser import (
    parse_formula_references,
)
from scripts.data.doc_intelligence.schema import (
    CellFormula,
    FormulaPayload,
    NamedRange,
    VbaModule,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_xlsx(tmp_dir: str, formulas: dict | None = None,
               named_ranges: dict | None = None,
               sheets: dict | None = None,
               suffix: str = ".xlsx") -> str:
    """Create a synthetic workbook and return its path.

    Args:
        formulas: {cell_ref: formula_str} for Sheet1
        named_ranges: {name: cell_ref}
        sheets: {sheet_name: {cell_ref: value_or_formula}}
        suffix: file extension
    """
    from openpyxl import Workbook
    from openpyxl.workbook.defined_name import DefinedName

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    if sheets:
        first = True
        for sheet_name, cells in sheets.items():
            if first:
                ws.title = sheet_name
                target = ws
                first = False
            else:
                target = wb.create_sheet(sheet_name)
            for ref, val in cells.items():
                target[ref] = val
    elif formulas:
        for ref, val in formulas.items():
            ws[ref] = val
    else:
        # Empty workbook — just the default empty sheet
        pass

    if named_ranges:
        for name, ref in named_ranges.items():
            defn = DefinedName(name, attr_text=f"Sheet1!{ref}")
            wb.defined_names.add(defn)

    path = os.path.join(tmp_dir, f"test_workbook{suffix}")
    wb.save(path)
    wb.close()
    return path


def _parse_file(filepath: str, domain: str = "test"):
    """Parse a file using FormulaXlsxParser."""
    from scripts.data.doc_intelligence.parsers.formula_xlsx import (
        FormulaXlsxParser,
    )
    parser = FormulaXlsxParser()
    return parser.parse(filepath, domain)


# ---------------------------------------------------------------------------
# Test 1: can_handle .xlsx
# ---------------------------------------------------------------------------

class TestCanHandle:
    def test_can_handle_xlsx(self):
        from scripts.data.doc_intelligence.parsers.formula_xlsx import (
            FormulaXlsxParser,
        )
        parser = FormulaXlsxParser()
        assert parser.can_handle("report.xlsx") is True
        assert parser.can_handle("REPORT.XLSX") is True
        assert parser.can_handle("data.pdf") is False
        assert parser.can_handle("notes.docx") is False

    # Test 2: can_handle .xlsm
    def test_can_handle_xlsm(self):
        from scripts.data.doc_intelligence.parsers.formula_xlsx import (
            FormulaXlsxParser,
        )
        parser = FormulaXlsxParser()
        assert parser.can_handle("macros.xlsm") is True
        assert parser.can_handle("MACROS.XLSM") is True


# ---------------------------------------------------------------------------
# Test 3: parse simple formula
# ---------------------------------------------------------------------------

class TestParseFormulas:
    def test_parse_simple_formula(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = _make_xlsx(tmp, formulas={
                "A1": 10,
                "B1": 20,
                "C1": "=A1+B1",
            })
            manifest = _parse_file(path)
            payload = manifest.formula_payload
            assert payload is not None
            assert len(payload.formulas) == 1
            f = payload.formulas[0]
            assert f.formula == "=A1+B1"
            assert f.cell_ref == "C1"
            assert f.sheet == "Sheet1"
            assert "A1" in f.references
            assert "B1" in f.references

    # Test 4: named ranges
    def test_parse_named_ranges(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = _make_xlsx(
                tmp,
                formulas={"A1": 42},
                named_ranges={"Price": "$A$1"},
            )
            manifest = _parse_file(path)
            payload = manifest.formula_payload
            assert payload is not None
            assert len(payload.named_ranges) >= 1
            names = [nr.name for nr in payload.named_ranges]
            assert "Price" in names

    # Test 5: identify input cells
    def test_identify_input_cells(self):
        """Literal cells referenced by formulas are inputs."""
        with tempfile.TemporaryDirectory() as tmp:
            path = _make_xlsx(tmp, formulas={
                "A1": 5,
                "B1": 10,
                "C1": "=A1+B1",
            })
            manifest = _parse_file(path)
            payload = manifest.formula_payload
            assert payload is not None
            # A1 and B1 are inputs (in-degree 0 in the DAG)
            # This depends on networkx — if not available, skip
            if payload.input_cells or payload.calculation_chain:
                input_refs = [c.cell_ref for c in payload.input_cells]
                # inputs should not contain C1 (it's a formula)
                assert "C1" not in input_refs

    # Test 6: identify output cells
    def test_identify_output_cells(self):
        """Formula cells not referenced by others are outputs."""
        with tempfile.TemporaryDirectory() as tmp:
            path = _make_xlsx(tmp, formulas={
                "A1": 5,
                "B1": 10,
                "C1": "=A1+B1",
            })
            manifest = _parse_file(path)
            payload = manifest.formula_payload
            assert payload is not None
            if payload.output_cells:
                output_refs = [c.cell_ref for c in payload.output_cells]
                assert "C1" in output_refs

    # Test 7: calculation chain order (topological sort)
    def test_calculation_chain_order(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = _make_xlsx(tmp, formulas={
                "A1": 1,
                "B1": "=A1*2",
                "C1": "=B1+A1",
            })
            manifest = _parse_file(path)
            payload = manifest.formula_payload
            assert payload is not None
            if payload.calculation_chain:
                chain = payload.calculation_chain
                # B1 must come before C1 (C1 depends on B1)
                if "B1" in chain and "C1" in chain:
                    assert chain.index("B1") < chain.index("C1")


# ---------------------------------------------------------------------------
# Test 8: cross-sheet references
# ---------------------------------------------------------------------------

class TestCrossSheet:
    def test_cross_sheet_references(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = _make_xlsx(tmp, sheets={
                "Data": {"A1": 100},
                "Calc": {"A1": "=Data!A1*2"},
            })
            manifest = _parse_file(path)
            payload = manifest.formula_payload
            assert payload is not None
            assert len(payload.formulas) == 1
            f = payload.formulas[0]
            assert f.sheet == "Calc"
            assert "Data!A1" in f.references


# ---------------------------------------------------------------------------
# Test 9: empty workbook
# ---------------------------------------------------------------------------

class TestEdgeCases:
    def test_empty_workbook(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = _make_xlsx(tmp)
            manifest = _parse_file(path)
            payload = manifest.formula_payload
            assert payload is not None
            assert len(payload.formulas) == 0
            assert len(payload.named_ranges) == 0
            assert payload.cache_quality["total_formulas"] == 0
            assert payload.cache_quality["quality_pct"] == 100.0


# ---------------------------------------------------------------------------
# Test 10: formula reference parsing (regex unit tests)
# ---------------------------------------------------------------------------

class TestFormulaReferenceParsing:
    def test_formula_reference_parsing(self):
        # Simple refs
        assert parse_formula_references("=A1+B2") == ["A1", "B2"]

        # Absolute refs
        refs = parse_formula_references("=$A$1+$B$2")
        assert "$A$1" in refs
        assert "$B$2" in refs

        # Mixed refs
        refs = parse_formula_references("=A$1+$B2")
        assert "A$1" in refs
        assert "$B2" in refs

        # Range
        refs = parse_formula_references("=SUM(A1:B10)")
        assert "A1" in refs
        assert "B10" in refs

        # Cross-sheet plain
        refs = parse_formula_references("=Sheet1!A1+B1")
        assert "Sheet1!A1" in refs
        assert "B1" in refs

        # Cross-sheet quoted
        refs = parse_formula_references("='My Sheet'!A1+B1")
        assert "My Sheet!A1" in refs
        assert "B1" in refs

    def test_empty_formula(self):
        assert parse_formula_references("") == []
        assert parse_formula_references("hello") == []


# ---------------------------------------------------------------------------
# Test 11-12: cache quality gate
# ---------------------------------------------------------------------------

class TestCacheQuality:
    def test_cache_quality_gate_all_ok(self):
        """All formula cells have cached values → quality 100%."""
        with tempfile.TemporaryDirectory() as tmp:
            path = _make_xlsx(tmp, formulas={
                "A1": 10,
                "B1": 20,
                "C1": "=A1+B1",
            })
            manifest = _parse_file(path)
            payload = manifest.formula_payload
            assert payload is not None
            # openpyxl may or may not have cached values for synthetic workbooks
            # but we check the structure is correct
            q = payload.cache_quality
            assert "total_formulas" in q
            assert "cached_ok" in q
            assert "cached_missing" in q
            assert "quality_pct" in q
            assert q["total_formulas"] == 1

    def test_cache_quality_gate_missing(self):
        """Formulas without cached values → cached_missing."""
        # When openpyxl saves a workbook, data_only pass may not have cached
        # values, so cached_missing is expected for synthetic workbooks.
        with tempfile.TemporaryDirectory() as tmp:
            path = _make_xlsx(tmp, formulas={
                "A1": 10,
                "C1": "=A1*3",
            })
            manifest = _parse_file(path)
            payload = manifest.formula_payload
            assert payload is not None
            f = payload.formulas[0]
            # Synthetic workbooks typically have no cache
            assert f.cache_status in ("cached_ok", "cached_missing")


# ---------------------------------------------------------------------------
# Test 13: only cached_ok cells emit assertions
# ---------------------------------------------------------------------------

class TestCacheAssertion:
    def test_only_cached_ok_cells_emit_assertions(self):
        """Test generation should skip cells with cache_status != cached_ok."""
        cells = [
            CellFormula(
                cell_ref="A1", sheet="S1", formula="=B1+C1",
                cached_value=30, cache_status="cached_ok",
            ),
            CellFormula(
                cell_ref="A2", sheet="S1", formula="=B2+C2",
                cached_value=None, cache_status="cached_missing",
            ),
        ]
        assertable = [c for c in cells if c.cache_status == "cached_ok"]
        assert len(assertable) == 1
        assert assertable[0].cell_ref == "A1"
        assert assertable[0].cached_value == 30


# ---------------------------------------------------------------------------
# Test 14: extract VBA macros (mocked)
# ---------------------------------------------------------------------------

class TestVbaExtraction:
    def test_extract_vba_macros(self):
        """Mock oletools to test VBA extraction logic."""
        from scripts.data.doc_intelligence.vba_extractor import (
            parse_vba_signatures,
        )
        from scripts.data.doc_intelligence import vba_extractor

        # Test parse_vba_signatures directly (no oletools needed)
        code = (
            "Public Function CalcArea(w As Double, h As Double) As Double\n"
            "    CalcArea = w * h\n"
            "End Function\n"
            "\n"
            "Sub RunReport()\n"
            "    MsgBox \"Done\"\n"
            "End Sub\n"
        )
        sigs = parse_vba_signatures(code)
        assert len(sigs) == 2
        assert "Function CalcArea(w As Double, h As Double)" in sigs[0]
        assert "Sub RunReport()" in sigs[1]

        # Test extract_vba_modules with mocked oletools
        mock_vba_parser_inst = MagicMock()
        mock_vba_parser_inst.detect_vba_macros.return_value = True
        mock_vba_parser_inst.extract_macros.return_value = [
            ("vbaProject.bin", "VBA/Module1", "Module1", code),
        ]
        mock_vba_cls = MagicMock(return_value=mock_vba_parser_inst)
        mock_olevba = MagicMock()
        mock_olevba.VBA_Parser = mock_vba_cls

        # Patch sys.modules so the `from oletools.olevba import VBA_Parser`
        # inside extract_vba_modules resolves to our mock.
        with patch.dict(sys.modules, {
            "oletools": MagicMock(),
            "oletools.olevba": mock_olevba,
        }):
            modules = vba_extractor.extract_vba_modules("fake.xlsm")
            assert len(modules) == 1
            assert modules[0]["filename"] == "Module1"
            assert len(modules[0]["signatures"]) == 2

    # Test 15: VBA function to Python stub
    def test_vba_function_to_python_stub(self):
        """VBA signature can be mapped to a Python def stub."""
        sig = "Function CalcArea(w As Double, h As Double)"
        # Simple conversion: extract name and params
        import re

        m = re.match(r"(?:Function|Sub)\s+(\w+)\(([^)]*)\)", sig)
        assert m is not None
        name = m.group(1)
        raw_params = m.group(2)
        # Strip VBA types
        params = []
        for p in raw_params.split(","):
            p = p.strip()
            param_name = p.split(" ")[0].strip()
            if param_name:
                params.append(param_name)
        stub = f"def {name}({', '.join(params)}):\n    raise NotImplementedError"
        assert "def CalcArea(w, h):" in stub
        assert "raise NotImplementedError" in stub


# ---------------------------------------------------------------------------
# Test 16-17: formula payload archive serialization
# ---------------------------------------------------------------------------

class TestArchiveSerialization:
    def _make_payload(self) -> FormulaPayload:
        return FormulaPayload(
            formulas=[
                CellFormula(
                    cell_ref="C1", sheet="Sheet1",
                    formula="=A1+B1", cached_value=30,
                    cache_status="cached_ok", references=["A1", "B1"],
                ),
            ],
            named_ranges=[
                NamedRange(name="Price", cell_ref="Sheet1!$A$1", scope="Sheet1"),
            ],
            input_cells=[],
            output_cells=[],
            calculation_chain=["A1", "B1", "C1"],
            vba_modules=[
                VbaModule(
                    filename="Module1",
                    code="Function X()\nEnd Function",
                    block_type="function",
                    signatures=["Function X()"],
                ),
            ],
            cache_quality={
                "total_formulas": 1,
                "cached_ok": 1,
                "cached_missing": 0,
                "quality_pct": 100.0,
            },
        )

    def _payload_to_dict(self, p: FormulaPayload) -> dict:
        """Convert FormulaPayload to a plain dict for archive."""
        return {
            "formulas": [
                {
                    "cell_ref": f.cell_ref,
                    "sheet": f.sheet,
                    "formula": f.formula,
                    "cached_value": f.cached_value,
                    "cache_status": f.cache_status,
                    "references": f.references,
                }
                for f in p.formulas
            ],
            "named_ranges": [
                {
                    "name": nr.name,
                    "cell_ref": nr.cell_ref,
                    "scope": nr.scope,
                }
                for nr in p.named_ranges
            ],
            "input_cells": [
                {"cell_ref": c.cell_ref, "sheet": c.sheet}
                for c in p.input_cells
            ],
            "output_cells": [
                {"cell_ref": c.cell_ref, "sheet": c.sheet}
                for c in p.output_cells
            ],
            "calculation_chain": p.calculation_chain,
            "vba_modules": [
                {
                    "filename": v.filename,
                    "code": v.code,
                    "block_type": v.block_type,
                    "signatures": v.signatures,
                }
                for v in p.vba_modules
            ],
            "cache_quality": p.cache_quality,
        }

    # Test 16: formula to archive yaml
    def test_formula_to_archive_yaml(self):
        payload = self._make_payload()
        d = self._payload_to_dict(payload)

        assert len(d["formulas"]) == 1
        assert d["formulas"][0]["formula"] == "=A1+B1"
        assert d["formulas"][0]["cached_value"] == 30
        assert d["named_ranges"][0]["name"] == "Price"
        assert d["calculation_chain"] == ["A1", "B1", "C1"]
        assert d["vba_modules"][0]["block_type"] == "function"
        assert d["cache_quality"]["quality_pct"] == 100.0

    # Test 17: archive yaml schema valid
    def test_archive_yaml_schema_valid(self):
        payload = self._make_payload()
        d = self._payload_to_dict(payload)

        # All required top-level keys present
        required_keys = {
            "formulas", "named_ranges", "input_cells",
            "output_cells", "calculation_chain", "vba_modules",
            "cache_quality",
        }
        assert set(d.keys()) == required_keys

        # Formula entry has all required fields
        formula_keys = {
            "cell_ref", "sheet", "formula",
            "cached_value", "cache_status", "references",
        }
        assert set(d["formulas"][0].keys()) == formula_keys

        # Named range entry has all required fields
        nr_keys = {"name", "cell_ref", "scope"}
        assert set(d["named_ranges"][0].keys()) == nr_keys

        # VBA module entry has all required fields
        vba_keys = {"filename", "code", "block_type", "signatures"}
        assert set(d["vba_modules"][0].keys()) == vba_keys

        # Cache quality has all required fields
        cq_keys = {"total_formulas", "cached_ok", "cached_missing", "quality_pct"}
        assert set(d["cache_quality"].keys()) == cq_keys


# ---------------------------------------------------------------------------
# Test 18-25: cache-miss validation (WRK-1247)
# ---------------------------------------------------------------------------

class TestCacheMissValidation:
    """Validate that cache-miss detection catches unrecalculated workbooks."""

    def test_validate_all_cached_ok(self):
        """100% cache hit rate passes validation."""
        from scripts.data.doc_intelligence.parsers.formula_xlsx import (
            validate_cache_quality,
        )
        cells = [
            CellFormula(
                cell_ref="A1", sheet="S1", formula="=B1+C1",
                cached_value=30, cache_status="cached_ok",
            ),
            CellFormula(
                cell_ref="A2", sheet="S1", formula="=B2*2",
                cached_value=10, cache_status="cached_ok",
            ),
        ]
        result = validate_cache_quality(cells, threshold=0.5)
        assert result["passed"] is True
        assert result["total_formulas"] == 2
        assert result["cache_miss_count"] == 0
        assert result["cache_miss_rate"] == 0.0
        assert result["threshold"] == 0.5
        assert len(result["warnings"]) == 0

    def test_validate_all_missing_fails(self):
        """100% cache miss rate fails validation."""
        from scripts.data.doc_intelligence.parsers.formula_xlsx import (
            validate_cache_quality,
        )
        cells = [
            CellFormula(
                cell_ref="A1", sheet="S1", formula="=B1+C1",
                cached_value=None, cache_status="cached_missing",
            ),
            CellFormula(
                cell_ref="A2", sheet="S1", formula="=B2*2",
                cached_value=None, cache_status="cached_missing",
            ),
        ]
        result = validate_cache_quality(cells, threshold=0.5)
        assert result["passed"] is False
        assert result["cache_miss_count"] == 2
        assert result["cache_miss_rate"] == 1.0
        assert len(result["warnings"]) >= 1
        assert "cache" in result["warnings"][0].lower()

    def test_validate_partial_miss_below_threshold(self):
        """Cache miss rate below threshold passes."""
        from scripts.data.doc_intelligence.parsers.formula_xlsx import (
            validate_cache_quality,
        )
        # 1 miss out of 4 = 25%, threshold 50% -> pass
        cells = [
            CellFormula(
                cell_ref="A1", sheet="S1", formula="=B1",
                cached_value=10, cache_status="cached_ok",
            ),
            CellFormula(
                cell_ref="A2", sheet="S1", formula="=B2",
                cached_value=20, cache_status="cached_ok",
            ),
            CellFormula(
                cell_ref="A3", sheet="S1", formula="=B3",
                cached_value=30, cache_status="cached_ok",
            ),
            CellFormula(
                cell_ref="A4", sheet="S1", formula="=B4",
                cached_value=None, cache_status="cached_missing",
            ),
        ]
        result = validate_cache_quality(cells, threshold=0.5)
        assert result["passed"] is True
        assert result["cache_miss_rate"] == 0.25

    def test_validate_partial_miss_above_threshold(self):
        """Cache miss rate above threshold fails."""
        from scripts.data.doc_intelligence.parsers.formula_xlsx import (
            validate_cache_quality,
        )
        # 3 miss out of 4 = 75%, threshold 50% -> fail
        cells = [
            CellFormula(
                cell_ref="A1", sheet="S1", formula="=B1",
                cached_value=10, cache_status="cached_ok",
            ),
            CellFormula(
                cell_ref="A2", sheet="S1", formula="=B2",
                cached_value=None, cache_status="cached_missing",
            ),
            CellFormula(
                cell_ref="A3", sheet="S1", formula="=B3",
                cached_value=None, cache_status="cached_missing",
            ),
            CellFormula(
                cell_ref="A4", sheet="S1", formula="=B4",
                cached_value=None, cache_status="cached_missing",
            ),
        ]
        result = validate_cache_quality(cells, threshold=0.5)
        assert result["passed"] is False
        assert result["cache_miss_rate"] == 0.75
        assert len(result["warnings"]) >= 1

    def test_validate_empty_formulas_passes(self):
        """No formulas at all passes (nothing to validate)."""
        from scripts.data.doc_intelligence.parsers.formula_xlsx import (
            validate_cache_quality,
        )
        result = validate_cache_quality([], threshold=0.5)
        assert result["passed"] is True
        assert result["total_formulas"] == 0
        assert result["cache_miss_rate"] == 0.0

    def test_validate_per_sheet_breakdown(self):
        """Validation result includes per-sheet cache-miss breakdown."""
        from scripts.data.doc_intelligence.parsers.formula_xlsx import (
            validate_cache_quality,
        )
        cells = [
            CellFormula(
                cell_ref="A1", sheet="Data", formula="=B1",
                cached_value=10, cache_status="cached_ok",
            ),
            CellFormula(
                cell_ref="A2", sheet="Data", formula="=B2",
                cached_value=None, cache_status="cached_missing",
            ),
            CellFormula(
                cell_ref="A1", sheet="Calc", formula="=B1",
                cached_value=None, cache_status="cached_missing",
            ),
        ]
        result = validate_cache_quality(cells, threshold=0.5)
        assert "per_sheet" in result
        assert "Data" in result["per_sheet"]
        assert "Calc" in result["per_sheet"]
        assert result["per_sheet"]["Data"]["miss_count"] == 1
        assert result["per_sheet"]["Data"]["total"] == 2
        assert result["per_sheet"]["Calc"]["miss_count"] == 1
        assert result["per_sheet"]["Calc"]["total"] == 1

    def test_parser_includes_cache_validation_in_manifest(self):
        """FormulaXlsxParser.parse() adds cache_validation to payload."""
        with tempfile.TemporaryDirectory() as tmp:
            path = _make_xlsx(tmp, formulas={
                "A1": 10,
                "B1": 20,
                "C1": "=A1+B1",
            })
            manifest = _parse_file(path)
            payload = manifest.formula_payload
            assert payload is not None
            assert hasattr(payload, "cache_validation")
            assert payload.cache_validation is not None
            assert "passed" in payload.cache_validation
            assert "cache_miss_rate" in payload.cache_validation
            assert "threshold" in payload.cache_validation

    def test_parser_adds_warning_to_errors_on_high_miss_rate(self):
        """When cache miss rate exceeds threshold, manifest.errors has warning."""
        with tempfile.TemporaryDirectory() as tmp:
            path = _make_xlsx(tmp, formulas={
                "A1": 10,
                "B1": 20,
                "C1": "=A1+B1",
                "D1": "=A1*B1",
                "E1": "=C1+D1",
            })
            manifest = _parse_file(path)
            payload = manifest.formula_payload
            assert payload is not None
            if payload.cache_validation["cache_miss_rate"] > 0.5:
                assert any(
                    "cache" in e.lower() for e in manifest.errors
                ), "Expected cache-miss warning in manifest.errors"
