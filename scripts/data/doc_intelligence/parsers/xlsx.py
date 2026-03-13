"""XLSX parser using openpyxl — each sheet becomes a table."""

import hashlib
from datetime import datetime, timezone
from pathlib import Path
from typing import List

from scripts.data.doc_intelligence.parsers.base import BaseParser
from scripts.data.doc_intelligence.schema import (
    DocumentManifest,
    DocumentMetadata,
    ExtractedTable,
    SourceLocation,
)


def _compute_checksum(filepath: str) -> str:
    sha = hashlib.sha256()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            sha.update(chunk)
    return sha.hexdigest()


class XlsxParser(BaseParser):
    """Extract each worksheet as a table from XLSX files."""

    def can_handle(self, filepath: str) -> bool:
        return Path(filepath).suffix.lower() == ".xlsx"

    def parse(self, filepath: str, domain: str) -> DocumentManifest:
        p = Path(filepath)
        meta = DocumentMetadata(
            filename=p.name,
            format="xlsx",
            size_bytes=p.stat().st_size,
            checksum=_compute_checksum(filepath),
            extraction_timestamp=datetime.now(timezone.utc).isoformat(),
        )
        tables: List[ExtractedTable] = []
        errors: List[str] = []

        try:
            from openpyxl import load_workbook

            wb = load_workbook(str(p), read_only=True, data_only=True)
            meta.sheets = len(wb.sheetnames)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                all_rows = []
                for row in ws.iter_rows(values_only=True):
                    str_row = [str(c) if c is not None else "" for c in row]
                    if any(cell.strip() for cell in str_row):
                        all_rows.append(str_row)

                if not all_rows:
                    continue

                header = all_rows[0]
                data_rows = all_rows[1:]
                tables.append(
                    ExtractedTable(
                        title=sheet_name,
                        columns=header,
                        rows=data_rows,
                        source=SourceLocation(
                            document=p.name, sheet=sheet_name
                        ),
                    )
                )
            wb.close()
        except Exception as exc:
            errors.append(f"XLSX extraction failed: {exc}")

        return DocumentManifest(
            version="1.0.0",
            tool="extract-document/1.0.0",
            domain=domain,
            metadata=meta,
            sections=[],
            tables=tables,
            figure_refs=[],
            extraction_stats={
                "sections": 0,
                "tables": len(tables),
                "figure_refs": 0,
            },
            errors=errors,
        )
