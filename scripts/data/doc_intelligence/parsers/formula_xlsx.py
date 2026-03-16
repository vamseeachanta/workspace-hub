"""FormulaXlsxParser — dual-pass XLSX/XLSM loader for formula extraction.

Pass 1 (data_only=True): captures cached computed values.
Pass 2 (data_only=False): captures formula strings.
Merges both into a FormulaPayload attached to the DocumentManifest.
"""

import hashlib
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

from scripts.data.doc_intelligence.formula_reference_parser import (
    parse_formula_references,
)
from scripts.data.doc_intelligence.parsers.base import BaseParser
from scripts.data.doc_intelligence.schema import (
    CellFormula,
    DocumentManifest,
    DocumentMetadata,
    ExtractedTable,
    FormulaPayload,
    NamedRange,
    SourceLocation,
    VbaModule,
)

_EXTENSIONS = {".xlsx", ".xlsm"}

# Default threshold: if >50% of formula cells have no cached value,
# the workbook was likely never recalculated (e.g., LibreOffice export).
_DEFAULT_CACHE_MISS_THRESHOLD = 0.5


def validate_cache_quality(
    formula_cells: List[CellFormula],
    threshold: float = _DEFAULT_CACHE_MISS_THRESHOLD,
) -> Dict[str, Any]:
    """Check whether cached-value miss rate exceeds the threshold.

    A workbook saved without recalculation (e.g., LibreOffice export)
    will have None for ALL formula cells via data_only=True, causing
    silent test failures when cached values are used as assertions.

    Args:
        formula_cells: List of CellFormula from the dual-pass extraction.
        threshold: Maximum acceptable cache-miss rate (0.0-1.0).

    Returns:
        Dict with keys: passed, total_formulas, cache_miss_count,
        cache_miss_rate, threshold, warnings, per_sheet.
    """
    total = len(formula_cells)
    if total == 0:
        return {
            "passed": True,
            "total_formulas": 0,
            "cache_miss_count": 0,
            "cache_miss_rate": 0.0,
            "threshold": threshold,
            "warnings": [],
            "per_sheet": {},
        }

    miss_count = sum(
        1 for c in formula_cells if c.cache_status == "cached_missing"
    )
    miss_rate = miss_count / total

    # Per-sheet breakdown
    per_sheet: Dict[str, Dict[str, Any]] = {}
    for cell in formula_cells:
        sheet = cell.sheet
        if sheet not in per_sheet:
            per_sheet[sheet] = {"total": 0, "miss_count": 0}
        per_sheet[sheet]["total"] += 1
        if cell.cache_status == "cached_missing":
            per_sheet[sheet]["miss_count"] += 1

    for sheet_stats in per_sheet.values():
        t = sheet_stats["total"]
        sheet_stats["miss_rate"] = (
            round(sheet_stats["miss_count"] / t, 4) if t else 0.0
        )

    passed = miss_rate <= threshold
    warnings: List[str] = []
    if not passed:
        warnings.append(
            f"Cache-miss validation failed: {miss_count}/{total} "
            f"formula cells ({miss_rate:.0%}) have no cached value. "
            f"Threshold is {threshold:.0%}. The workbook may not have "
            f"been recalculated before saving — cached values are "
            f"unreliable for test assertions."
        )

    return {
        "passed": passed,
        "total_formulas": total,
        "cache_miss_count": miss_count,
        "cache_miss_rate": round(miss_rate, 4),
        "threshold": threshold,
        "warnings": warnings,
        "per_sheet": per_sheet,
    }


def _compute_checksum(filepath: str) -> str:
    sha = hashlib.sha256()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            sha.update(chunk)
    return sha.hexdigest()


def _cell_ref_str(col: int, row: int) -> str:
    """Convert 1-based column and row to Excel ref like 'A1'."""
    result = ""
    c = col
    while c > 0:
        c, remainder = divmod(c - 1, 26)
        result = chr(65 + remainder) + result
    return f"{result}{row}"


class FormulaXlsxParser(BaseParser):
    """Extract formulas, named ranges, and cached values from XLSX/XLSM."""

    def can_handle(self, filepath: str) -> bool:
        return Path(filepath).suffix.lower() in _EXTENSIONS

    def parse(self, filepath: str, domain: str) -> DocumentManifest:
        p = Path(filepath)
        meta = DocumentMetadata(
            filename=p.name,
            format=p.suffix.lstrip(".").lower(),
            size_bytes=p.stat().st_size,
            checksum=_compute_checksum(filepath),
            extraction_timestamp=datetime.now(timezone.utc).isoformat(),
        )

        from openpyxl import load_workbook

        # Pass 1: cached values
        wb_data = load_workbook(str(p), read_only=True, data_only=True)
        cached_values: Dict[str, Any] = {}  # "Sheet!CellRef" -> value
        tables: List[ExtractedTable] = []

        meta.sheets = len(wb_data.sheetnames)

        for sheet_name in wb_data.sheetnames:
            ws = wb_data[sheet_name]
            # Skip chart-only sheets (Chartsheet has no iter_rows)
            if not hasattr(ws, "iter_rows"):
                continue
            all_rows: List[List[str]] = []
            for row in ws.iter_rows(values_only=False):
                str_row: List[str] = []
                for cell in row:
                    val = cell.value
                    if val is not None:
                        key = f"{sheet_name}!{_cell_ref_str(cell.column, cell.row)}"
                        cached_values[key] = val
                    str_row.append(str(val) if val is not None else "")
                if any(c.strip() for c in str_row):
                    all_rows.append(str_row)

            if all_rows:
                tables.append(
                    ExtractedTable(
                        title=sheet_name,
                        columns=all_rows[0],
                        rows=all_rows[1:],
                        source=SourceLocation(
                            document=p.name, sheet=sheet_name
                        ),
                    )
                )
        wb_data.close()

        # Pass 2: formulas
        wb_formula = load_workbook(str(p), read_only=False, data_only=False)
        formula_cells: List[CellFormula] = []

        for sheet_name in wb_formula.sheetnames:
            ws = wb_formula[sheet_name]
            if not hasattr(ws, "iter_rows"):
                continue
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if isinstance(val, str) and val.startswith("="):
                        ref = _cell_ref_str(cell.column, cell.row)
                        full_key = f"{sheet_name}!{ref}"
                        cached = cached_values.get(full_key)
                        refs = parse_formula_references(val)
                        status = "cached_ok" if cached is not None else "cached_missing"
                        formula_cells.append(
                            CellFormula(
                                cell_ref=ref,
                                sheet=sheet_name,
                                formula=val,
                                cached_value=cached,
                                cache_status=status,
                                references=refs,
                            )
                        )

        # Named ranges
        named_ranges: List[NamedRange] = []
        for name, defn in wb_formula.defined_names.items():
            try:
                for title, coord in defn.destinations:
                    named_ranges.append(
                        NamedRange(
                            name=name,
                            cell_ref=f"{title}!{coord}" if title else coord,
                            scope=title,
                        )
                    )
            except Exception:
                pass  # skip malformed definitions
        wb_formula.close()

        # Classify cells using chain builder (soft dependency)
        inputs: List[CellFormula] = []
        outputs: List[CellFormula] = []
        chain: List[str] = []
        max_graph_cells = 50_000  # Skip graph classification for very large files
        if len(formula_cells) <= max_graph_cells:
            try:
                from scripts.data.doc_intelligence.formula_chain_builder import (
                    build_dependency_graph,
                    classify_cells,
                )

                g = build_dependency_graph(formula_cells)
                classification = classify_cells(g)
                chain = classification["chain"]
                input_refs = set(classification["inputs"])
                output_refs = set(classification["outputs"])

                cell_map = {c.cell_ref: c for c in formula_cells}
                for ref in input_refs:
                    if ref in cell_map:
                        inputs.append(cell_map[ref])
                for ref in output_refs:
                    if ref in cell_map:
                        outputs.append(cell_map[ref])
            except Exception as _classify_exc:
                import sys
                print(f"CLASSIFY ERROR ({len(formula_cells)} cells): {_classify_exc}", file=sys.stderr)

        # VBA extraction (soft dependency, .xlsm only)
        vba_modules: List[VbaModule] = []
        if p.suffix.lower() == ".xlsm":
            try:
                from scripts.data.doc_intelligence.vba_extractor import (
                    extract_vba_modules,
                )

                raw = extract_vba_modules(str(p))
                for mod in raw:
                    vba_modules.append(
                        VbaModule(
                            filename=mod["filename"],
                            code=mod["code"],
                            block_type=mod["block_type"],
                            signatures=mod.get("signatures", []),
                        )
                    )
            except ImportError:
                pass

        # Cache quality
        total = len(formula_cells)
        ok = sum(1 for c in formula_cells if c.cache_status == "cached_ok")
        cache_quality: Dict[str, Any] = {
            "total_formulas": total,
            "cached_ok": ok,
            "cached_missing": total - ok,
            "quality_pct": round(ok / total * 100, 1) if total else 100.0,
        }

        # Cache-miss validation (WRK-1247)
        cache_validation = validate_cache_quality(formula_cells)

        payload = FormulaPayload(
            formulas=formula_cells,
            named_ranges=named_ranges,
            input_cells=inputs,
            output_cells=outputs,
            calculation_chain=chain,
            vba_modules=vba_modules,
            cache_quality=cache_quality,
            cache_validation=cache_validation,
        )

        errors: List[str] = []
        errors.extend(cache_validation.get("warnings", []))
        return DocumentManifest(
            version="1.0.0",
            tool="formula-extract/1.0.0",
            domain=domain,
            metadata=meta,
            sections=[],
            tables=tables,
            figure_refs=[],
            extraction_stats={
                "sections": 0,
                "tables": len(tables),
                "figure_refs": 0,
                "formulas": total,
                "named_ranges": len(named_ranges),
            },
            errors=errors,
            formula_payload=payload,
        )
