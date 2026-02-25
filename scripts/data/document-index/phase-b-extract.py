#!/usr/bin/env python3
# ABOUTME: Phase B — Text extraction and LLM summarisation for indexed documents (WRK-309)
# ABOUTME: Resume-safe batch processor: reads index.jsonl, writes summaries/<sha256>.json

"""
Usage:
    python phase-b-extract.py [--config config.yaml] [--limit N] [--source SRC] [--no-llm]
"""

import argparse
import hashlib
import json
import logging
import sqlite3
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import yaml

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

SCRIPT_DIR = Path(__file__).resolve().parent
HUB_ROOT = SCRIPT_DIR.parents[2]
DEFAULT_CONFIG = SCRIPT_DIR / "config.yaml"


def load_config(config_path: Path) -> Dict[str, Any]:
    """Load pipeline configuration from YAML."""
    with open(config_path) as f:
        return yaml.safe_load(f)


def load_index(index_path: Path) -> List[Dict]:
    """Load all records from index.jsonl."""
    records: List[Dict] = []
    if not index_path.exists():
        logger.error("Index not found: %s", index_path)
        return records
    with open(index_path) as f:
        for line in f:
            line = line.strip()
            if line:
                try:
                    records.append(json.loads(line))
                except json.JSONDecodeError:
                    continue
    return records


def summary_key_for(record: Dict) -> str:
    """Derive summary filename key from content hash or path hash."""
    key = record.get("content_hash") or record.get("path", "")
    return hashlib.sha256(key.encode()).hexdigest()[:16]


def extract_og_sqlite(record: Dict, cfg: Dict) -> Optional[str]:
    """Extract text from og_standards via SQLite document_text table."""
    db_path = cfg["sources"]["og_standards"]["db_path"]
    og_id = record.get("og_db_id")
    if not og_id or not Path(db_path).exists():
        return None
    conn = sqlite3.connect(db_path, timeout=30)
    try:
        cursor = conn.execute(
            "SELECT full_text FROM document_text WHERE document_id = ?",
            (og_id,),
        )
        row = cursor.fetchone()
        return row[0] if row else None
    finally:
        conn.close()


def extract_pdf(file_path: Path, cfg: Dict) -> tuple:
    """Extract text from PDF. Returns (text, method)."""
    size_mb = file_path.stat().st_size / (1024 * 1024)
    ext_cfg = cfg.get("extraction", {})
    threshold = ext_cfg.get("large_pdf_threshold_mb", 100)
    max_chars = ext_cfg.get("max_text_chars", 50000)

    if size_mb >= threshold:
        cli_path = HUB_ROOT / ext_cfg.get("large_pdf_cli", "")
        if cli_path.exists():
            try:
                result = subprocess.run(
                    [sys.executable, str(cli_path), str(file_path)],
                    capture_output=True, text=True, timeout=600,
                )
                if result.returncode == 0 and result.stdout.strip():
                    return result.stdout[:max_chars], "pdf_large_reader"
            except (subprocess.TimeoutExpired, OSError) as e:
                logger.warning("pdf-large-reader failed: %s", e)

    try:
        result = subprocess.run(
            ["pdftotext", "-q", str(file_path), "-"],
            capture_output=True, text=True, timeout=120,
        )
        if result.returncode == 0:
            return result.stdout[:max_chars], "pdftotext"
    except (subprocess.TimeoutExpired, FileNotFoundError) as e:
        logger.warning("pdftotext failed for %s: %s", file_path, e)
    return None, "pdftotext"


def extract_docx(file_path: Path, max_chars: int) -> Optional[str]:
    """Extract text from DOCX using python-docx."""
    try:
        from docx import Document
        doc = Document(str(file_path))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n".join(paragraphs)[:max_chars]
    except Exception as e:
        logger.warning("DOCX extraction failed for %s: %s", file_path, e)
        return None


def extract_xlsx(file_path: Path, max_rows: int) -> Optional[str]:
    """Extract headers + first N data rows from XLSX."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(file_path), read_only=True, data_only=True)
        parts: List[str] = []
        for sheet_name in wb.sheetnames[:3]:
            ws = wb[sheet_name]
            parts.append(f"Sheet: {sheet_name}")
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i > max_rows:
                    break
                row_str = " | ".join(str(c) if c is not None else "" for c in row)
                parts.append(row_str)
        wb.close()
        return "\n".join(parts) if parts else None
    except Exception as e:
        logger.warning("XLSX extraction failed for %s: %s", file_path, e)
        return None


def extract_text(record: Dict, cfg: Dict) -> tuple:
    """Route extraction by source and extension. Returns (text, method)."""
    source = record.get("source", "")
    ext = record.get("ext", "").lower()
    path_str = record.get("path", "")

    if source == "api_metadata":
        return None, "api_metadata"
    if record.get("is_cad"):
        return None, "skipped"
    if source == "og_standards":
        text = extract_og_sqlite(record, cfg)
        if text:
            return text, "og_sqlite"

    fpath = Path(path_str)
    if not fpath.exists():
        return None, "file_not_found"

    ext_cfg = cfg.get("extraction", {})
    max_chars = ext_cfg.get("max_text_chars", 50000)
    max_rows = ext_cfg.get("xlsx_max_rows", 5)
    md_chars = ext_cfg.get("md_max_chars", 2000)

    if ext == "pdf":
        return extract_pdf(fpath, cfg)
    elif ext == "docx":
        text = extract_docx(fpath, max_chars)
        return text, "docx"
    elif ext in ("xlsx", "xls"):
        text = extract_xlsx(fpath, max_rows)
        return text, "xlsx"
    elif ext in ("md", "txt", "yaml", "yml", "csv"):
        limit = md_chars if ext in ("md", "yaml", "yml") else max_chars
        try:
            text = fpath.read_text(errors="replace")[:limit]
            return text, "direct"
        except OSError:
            return None, "direct"
    return None, "unsupported"


def summarise_with_llm(text: str, cfg: Dict, daily_spend: float) -> tuple:
    """Call Anthropic API for summarisation. Returns (result_dict, cost)."""
    llm_cfg = cfg.get("llm", {})
    model = llm_cfg.get("model", "claude-haiku-4-5-20251001")
    budget = llm_cfg.get("daily_budget_usd", 20.0)

    if daily_spend >= budget:
        logger.warning("Daily LLM budget exhausted ($%.2f)", daily_spend)
        return None, 0.0

    try:
        import anthropic
        client = anthropic.Anthropic()
        prompt = (
            "Summarize this engineering document in 3-5 sentences. "
            "Extract: title, key topics, applicable standards mentioned, "
            "engineering discipline."
        )
        message = client.messages.create(
            model=model, max_tokens=500,
            messages=[{"role": "user", "content": f"{prompt}\n\n{text[:8000]}"}],
        )
        summary_text = message.content[0].text
        input_tok = message.usage.input_tokens
        output_tok = message.usage.output_tokens
        cost = (input_tok * 0.25 + output_tok * 1.25) / 1_000_000
        return {"summary": summary_text}, cost
    except Exception as exc:
        logger.warning("LLM summarisation failed: %s", exc)
        return None, 0.0


def write_summary(
    summaries_dir: Path, key: str, record: Dict,
    text: Optional[str], method: str, llm_result: Optional[Dict],
) -> None:
    """Write summary JSON for a single document."""
    summaries_dir.mkdir(parents=True, exist_ok=True)
    word_count = len(text.split()) if text else 0
    output = {
        "path": record.get("path"),
        "sha256": key,
        "title": (llm_result or {}).get("title") or record.get("doc_number")
                 or Path(record.get("path", "")).name,
        "summary": (llm_result or {}).get("summary"),
        "keywords": [],
        "page_count": None,
        "word_count": word_count,
        "extraction_method": method,
        "extracted_at": datetime.now().isoformat(timespec="seconds"),
    }
    out_path = summaries_dir / f"{key}.json"
    with open(out_path, "w") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Phase B: Text extraction and LLM summarisation (WRK-309)"
    )
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG)
    parser.add_argument("--limit", type=int, default=0, help="Max docs to process")
    parser.add_argument("--source", help="Process only this source type")
    parser.add_argument("--no-llm", action="store_true", help="Extract only, skip LLM")
    args = parser.parse_args()

    cfg = load_config(args.config)
    index_path = HUB_ROOT / cfg["output"]["index_path"]
    summaries_dir = HUB_ROOT / cfg["output"]["summaries_dir"]
    llm_cfg = cfg.get("llm", {})
    skip_below = llm_cfg.get("skip_below_words", 100)
    batch_size = llm_cfg.get("batch_size", 50)

    records = load_index(index_path)
    logger.info("Loaded %d index records", len(records))

    daily_spend = 0.0
    processed = 0
    skipped_existing = 0

    for rec in records:
        if args.source and rec.get("source") != args.source:
            continue

        key = summary_key_for(rec)
        sfile = summaries_dir / f"{key}.json"
        if sfile.exists():
            skipped_existing += 1
            continue

        if args.limit and processed >= args.limit:
            break

        text, method = extract_text(rec, cfg)

        if method in ("skipped", "api_metadata"):
            write_summary(summaries_dir, key, rec, None, method, None)
            processed += 1
            continue

        llm_result = None
        if not args.no_llm and text:
            word_count = len(text.split())
            if word_count >= skip_below:
                llm_result, cost = summarise_with_llm(text, cfg, daily_spend)
                daily_spend += cost

        write_summary(summaries_dir, key, rec, text, method, llm_result)
        processed += 1

        if processed % 100 == 0:
            logger.info("Processed %d, LLM spend: $%.2f", processed, daily_spend)
        if processed % batch_size == 0 and not args.no_llm:
            time.sleep(0.5)

    logger.info(
        "Phase B complete: %d processed, %d skipped (existing), LLM: $%.2f",
        processed, skipped_existing, daily_spend,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
