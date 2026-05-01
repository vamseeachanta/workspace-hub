#!/usr/bin/env python3
"""First-pass deterministic extraction for Elements high-value corpora."""
from __future__ import annotations

import json
from pathlib import Path
import openpyxl

OUT = Path('.planning/intel/elements-deep-extraction')
OUT.mkdir(parents=True, exist_ok=True)
(OUT / 'workbooks').mkdir(exist_ok=True)
(OUT / 'gis').mkdir(exist_ok=True)

WORKBOOKS = [
    Path('/mnt/ace/digitalmodel/references/suction-pile-sizing/SUCPILE.xlsm'),
    Path('/mnt/ace/digitalmodel/references/suction-pile-sizing/pile soil defl i doris system-suction pile.xlsm'),
    Path('/mnt/ace/digitalmodel/references/riser-toolbox/orcaflex tools/weibul statistics/Weibull Fitting-PR1-1in-1000yrWave-135deg-R3.xlsm'),
    Path('/mnt/ace/digitalmodel/references/riser-toolbox/orcaflex tools/gumbel statistics/_PR1-11in-EnvelopeExtraction.xlsm'),
]


def inspect_workbook(path: Path) -> dict:
    rec = {"path": str(path), "name": path.name, "size": path.stat().st_size, "sheets": []}
    try:
        wb = openpyxl.load_workbook(path, read_only=False, data_only=False, keep_vba=True)
        for ws in wb.worksheets:
            dims = {
                "title": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "non_empty_cells": 0,
                "formula_cells": 0,
                "labels": [],
                "formulas_sample": [],
            }
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    val = str(cell.value).strip()
                    if not val:
                        continue
                    dims["non_empty_cells"] += 1
                    if val.startswith('='):
                        dims["formula_cells"] += 1
                        if len(dims["formulas_sample"]) < 20:
                            dims["formulas_sample"].append({"cell": cell.coordinate, "value": val[:240]})
                    elif len(dims["labels"]) < 80:
                        dims["labels"].append({"cell": cell.coordinate, "value": val[:240]})
            rec["sheets"].append(dims)
    except Exception as exc:  # noqa: BLE001
        rec["error"] = repr(exc)
    return rec


def main() -> int:
    wb_summaries = []
    for path in WORKBOOKS:
        rec = inspect_workbook(path)
        (OUT / 'workbooks' / f'{path.name}.json').write_text(json.dumps(rec, indent=2), encoding='utf-8')
        wb_summaries.append({
            "path": str(path),
            "size": rec.get("size"),
            "error": rec.get("error"),
            "sheets": [
                {k: s[k] for k in ("title", "max_row", "max_column", "non_empty_cells", "formula_cells")}
                for s in rec.get("sheets", [])
            ],
        })
    (OUT / 'workbook-summary.json').write_text(json.dumps(wb_summaries, indent=2), encoding='utf-8')

    qgis = Path('/mnt/ace/digitalmodel/tools/qgis')
    qgis_meta = []
    for p in sorted(qgis.glob('*')):
        if p.is_file() and p.name != 'MOVE-LOG.md':
            qgis_meta.append({"path": str(p), "name": p.name, "size": p.stat().st_size, "suffix": p.suffix.lower()})
    (OUT / 'gis' / 'qgis-files.json').write_text(json.dumps(qgis_meta, indent=2), encoding='utf-8')

    print(json.dumps({"workbooks": wb_summaries, "qgis_files": qgis_meta}, indent=2))
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
